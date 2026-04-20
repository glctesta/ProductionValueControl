import logging
from pathlib import Path
from typing import Dict, Optional

from openpyxl import load_workbook

logger = logging.getLogger(__name__)


class ExcelService:
    """
    Legge il file Excel piu' recente nella directory configurata (per convenzione
    e' sempre lo stesso file, si considera quello con mtime piu' grande) e
    costruisce la mappa OrderNumber -> UnitPrice.

    Usa una cache in memoria invalidata dal mtime/path del file.
    """

    def __init__(self, directory: str, sheet_name: str):
        self.directory = directory
        self.sheet_name = sheet_name
        self._cache: Optional[Dict[str, float]] = None
        self._cache_mtime: Optional[float] = None
        self._cache_path: Optional[Path] = None

    def _latest_file(self) -> Path:
        root = Path(self.directory)
        if not root.exists():
            raise FileNotFoundError(f"Directory Excel non trovata: {self.directory}")
        candidates = [p for p in root.glob('*.xls*') if p.is_file() and not p.name.startswith('~$')]
        if not candidates:
            raise FileNotFoundError(f"Nessun file Excel in {self.directory}")
        candidates.sort(key=lambda p: p.stat().st_mtime, reverse=True)
        return candidates[0]

    def load_price_map(self) -> Dict[str, float]:
        path = self._latest_file()
        mtime = path.stat().st_mtime

        if (
            self._cache is not None
            and self._cache_path == path
            and self._cache_mtime == mtime
        ):
            return self._cache

        logger.info(f"Lettura Excel: {path.name} (mtime={mtime})")
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            if self.sheet_name not in wb.sheetnames:
                raise KeyError(
                    f"Foglio '{self.sheet_name}' non trovato in {path.name}. "
                    f"Disponibili: {wb.sheetnames}"
                )
            ws = wb[self.sheet_name]

            price_map: Dict[str, float] = {}
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row:
                    continue
                order_number = row[0] if len(row) > 0 else None
                unit_price = row[10] if len(row) > 10 else None

                if order_number is None:
                    continue
                order_str = str(order_number).strip()
                if not order_str:
                    continue

                try:
                    price_val = float(unit_price) if unit_price is not None else 0.0
                except (ValueError, TypeError):
                    continue

                price_map[order_str] = price_val
        finally:
            wb.close()

        logger.info(f"Mappa prezzi caricata: {len(price_map)} ordini")
        self._cache = price_map
        self._cache_path = path
        self._cache_mtime = mtime
        return price_map

    def source_file_name(self) -> str:
        try:
            return self._latest_file().name
        except Exception:
            return ""
