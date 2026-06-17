import logging
from datetime import date, datetime, timedelta
from io import BytesIO
from itertools import groupby
from typing import Dict, List, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from services.metrics_service import MetricsService

logger = logging.getLogger(__name__)


_IT_MONTHS = [
    "Gennaio", "Febbraio", "Marzo", "Aprile", "Maggio", "Giugno",
    "Luglio", "Agosto", "Settembre", "Ottobre", "Novembre", "Dicembre",
]
_IT_WEEKDAYS = ["Lun", "Mar", "Mer", "Gio", "Ven", "Sab", "Dom"]

# Palette
COL_TITLE = '1F4E78'
COL_HEADER = '2E75B6'
COL_WEEK = '305496'
COL_DAY = '8FAADC'
COL_KPI_LABEL = 'D9E1F2'
COL_KPI_VALUE = 'FFFFFF'
COL_KPI_BAND = '1F4E78'
COL_SUBTOTAL = 'DCE6F2'
COL_TOTAL = 'FFD966'
COL_NOTE_BG = 'FFF2CC'
COL_NOTE_BORDER = 'BF8F00'

CURRENCY_FMT = '#,##0.00 \u20AC'
INT_FMT = '#,##0'

THIN = Side(style='thin', color='BFBFBF')
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _a_center(wrap=False):
    return Alignment(horizontal='center', vertical='center', wrap_text=wrap)


def _a_right():
    return Alignment(horizontal='right', vertical='center')


def _a_left(wrap=True, indent=0):
    return Alignment(horizontal='left', vertical='center', wrap_text=wrap, indent=indent)


class ExportService:
    """
    Genera il workbook Excel con:
      - Tab 'Current Month': KPI in testa + dettaglio produzione del mese
        raggruppato per settimana ISO e per giorno (outline a 2 livelli).
        I semilavorati (product code che inizia per 'SL') sono esclusi e
        riportati nel tab dedicato.
      - Tab 'Semilavorati' (solo se presenti): dettaglio ordini SL con
        prezzo di SOLA MANODOPERA dal vecchio ERP.
    """

    LAST_COL = 8  # A..H

    def __init__(self, excel_service, sql_service, metrics_service):
        self.excel = excel_service
        self.sql = sql_service
        self.metrics = metrics_service

    # ---------------------------------------------------------------- Public
    def build_current_month_workbook(self, now: datetime) -> BytesIO:
        prod_day = MetricsService.current_production_date(now)
        year, month = prod_day.year, prod_day.month

        metrics = self.metrics.compute(now)

        month_start_dt = datetime(year, month, 1, 7, 30, 0)
        today_start = datetime(prod_day.year, prod_day.month, prod_day.day, 7, 30, 0)
        cutoff_dt = today_start + timedelta(days=1)

        rows = self.sql.get_month_production(month_start_dt, cutoff_dt)
        unique_orders = sorted({o for o, _, _ in rows})

        price_map = dict(self.excel.load_price_map())
        for o in unique_orders:
            if o not in price_map:
                fb = self.sql.get_price_from_resetservices(o)
                price_map[o] = fb if fb is not None else 0.0

        details = self.sql.get_orders_details_bulk(unique_orders)

        def product_code(order: str) -> str:
            return (details.get(order, {}).get('productCode') or '').upper()

        sl_orders = {o for o in unique_orders if product_code(o).startswith('SL')}
        # Nel tab principale manteniamo TUTTI gli ordini (inclusi SL) per non
        # nascondere produzione; gli SL verranno evidenziati in grassetto+corsivo.
        sl_rows = [(o, d, q) for o, d, q in rows if o in sl_orders]

        wb = Workbook()
        ws_main = wb.active
        ws_main.title = 'Current Month'

        self._build_main_sheet(
            ws=ws_main,
            now=now,
            prod_day=prod_day,
            year=year,
            month=month,
            metrics=metrics,
            rows=rows,
            details=details,
            price_map=price_map,
            sl_orders=sl_orders,
        )

        if sl_orders:
            labor_price_map: Dict[str, float] = {}
            for order in sl_orders:
                code = product_code(order)
                lp = self.sql.get_semilavorato_labor_price(code)
                labor_price_map[order] = float(lp) if lp is not None else 0.0

            ws_sl = wb.create_sheet('Semilavorati')
            self._build_semilavorati_sheet(
                ws=ws_sl,
                now=now,
                prod_day=prod_day,
                year=year,
                month=month,
                rows=sl_rows,
                details=details,
                labor_price_map=labor_price_map,
            )

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    def build_wip_workbook(self, wip_summary: List[dict], wip_details: List[dict], now: datetime) -> BytesIO:
        wb = Workbook()
        ws_main = wb.active
        ws_main.title = 'Sintesi WIP'

        # 1) Title + Subtitle
        ws_main.merge_cells(start_row=1, start_column=1, end_row=1, end_column=11)
        c_title = ws_main.cell(row=1, column=1, value="Sintesi Work In Progress (WIP)")
        c_title.font = Font(name='Calibri', size=15, bold=True, color='FFFFFF')
        c_title.fill = PatternFill('solid', fgColor=COL_TITLE)
        c_title.alignment = _a_center()
        ws_main.row_dimensions[1].height = 30

        ws_main.merge_cells(start_row=2, start_column=1, end_row=2, end_column=11)
        c_sub = ws_main.cell(row=2, column=1, value=f"Generato il {now.strftime('%d/%m/%Y %H:%M')} \u00b7 Stato attivo dei semilavorati in linea")
        c_sub.font = Font(name='Calibri', size=10, italic=True, color='595959')
        c_sub.alignment = Alignment(horizontal='right', vertical='center')

        # 2) Note
        self._write_wip_note(
            ws_main,
            start_row=4,
            text=(
                "Nota: Il WIP (Work In Progress) rappresenta il valore dei semilavorati in transito nella linea di produzione, "
                "scansionati a partire dalla fase di PTH ma non ancora completati (fase finale 142). Lo stato 'FAIL' indica che "
                "l'ultima scansione registrata per la scheda ha dato esito fallito. 'OK' indica che l'ultima scansione e' corretta."
            )
        )

        # 3) Headers
        header_row = 7
        headers = [
            'Order Number', 'Product Code', 'Product Name', 'Target Qty',
            'Qty WIP OK', 'Qty WIP FAIL', 'Total WIP Qty',
            'Prezzo Unit.', 'Valore WIP OK', 'Valore WIP FAIL', 'Valore WIP Totale'
        ]
        
        # Scrive gli header
        for idx, h in enumerate(headers, start=1):
            c = ws_main.cell(row=header_row, column=idx, value=h)
            c.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
            c.fill = PatternFill('solid', fgColor=COL_HEADER)
            c.alignment = _a_center(wrap=True)
            c.border = BORDER_ALL
        ws_main.row_dimensions[header_row].height = 30
        ws_main.freeze_panes = ws_main.cell(row=header_row + 1, column=1).coordinate

        # 4) Write Summary Rows
        cur = header_row + 1
        grand_ok = 0
        grand_fail = 0
        grand_total_qty = 0
        grand_val_ok = 0.0
        grand_val_fail = 0.0
        grand_val_total = 0.0

        for s in wip_summary:
            qty_ok = int(s.get('QtyOK', 0))
            qty_fail = int(s.get('QtyFAIL', 0))
            total_qty = qty_ok + qty_fail
            price = float(s.get('UnitPrice', 0.0))
            val_ok = qty_ok * price
            val_fail = qty_fail * price
            val_total = total_qty * price

            # Scrive i valori
            ws_main.cell(row=cur, column=1, value=s.get('OrderNumber')).alignment = _a_center()
            ws_main.cell(row=cur, column=2, value=s.get('ProductCode')).alignment = _a_center()
            ws_main.cell(row=cur, column=3, value=s.get('ProductName')).alignment = _a_left(wrap=True)
            
            t_qty = s.get('TargetQty', 0)
            ws_main.cell(row=cur, column=4, value=int(t_qty) if t_qty is not None else 0).number_format = INT_FMT
            ws_main.cell(row=cur, column=4).alignment = _a_right()
            
            ws_main.cell(row=cur, column=5, value=qty_ok).number_format = INT_FMT
            ws_main.cell(row=cur, column=5).alignment = _a_right()
            
            ws_main.cell(row=cur, column=6, value=qty_fail).number_format = INT_FMT
            ws_main.cell(row=cur, column=6).alignment = _a_right()
            
            ws_main.cell(row=cur, column=7, value=total_qty).number_format = INT_FMT
            ws_main.cell(row=cur, column=7).alignment = _a_right()
            
            p_cell = ws_main.cell(row=cur, column=8, value=price)
            p_cell.number_format = CURRENCY_FMT
            p_cell.alignment = _a_right()

            vo_cell = ws_main.cell(row=cur, column=9, value=val_ok)
            vo_cell.number_format = CURRENCY_FMT
            vo_cell.alignment = _a_right()

            vf_cell = ws_main.cell(row=cur, column=10, value=val_fail)
            vf_cell.number_format = CURRENCY_FMT
            vf_cell.alignment = _a_right()

            vt_cell = ws_main.cell(row=cur, column=11, value=val_total)
            vt_cell.number_format = CURRENCY_FMT
            vt_cell.alignment = _a_right()

            for col in range(1, 12):
                cell = ws_main.cell(row=cur, column=col)
                cell.font = Font(name='Calibri', size=10)
                cell.border = BORDER_ALL

            grand_ok += qty_ok
            grand_fail += qty_fail
            grand_total_qty += total_qty
            grand_val_ok += val_ok
            grand_val_fail += val_fail
            grand_val_total += val_total
            cur += 1

        # Grand total
        ws_main.merge_cells(start_row=cur, start_column=1, end_row=cur, end_column=3)
        lbl_cell = ws_main.cell(row=cur, column=1, value='TOTALE WIP')
        lbl_cell.font = Font(name='Calibri', size=11, bold=True)
        lbl_cell.alignment = Alignment(horizontal='right', vertical='center', indent=1)
        
        for col in range(1, 4):
            ws_main.cell(row=cur, column=col).fill = PatternFill('solid', fgColor=COL_TOTAL)
            ws_main.cell(row=cur, column=col).border = BORDER_ALL

        ws_main.cell(row=cur, column=4).fill = PatternFill('solid', fgColor=COL_TOTAL)
        ws_main.cell(row=cur, column=4).border = BORDER_ALL

        totals = [
            (5, grand_ok, INT_FMT),
            (6, grand_fail, INT_FMT),
            (7, grand_total_qty, INT_FMT),
            (8, None, None),
            (9, grand_val_ok, CURRENCY_FMT),
            (10, grand_val_fail, CURRENCY_FMT),
            (11, grand_val_total, CURRENCY_FMT),
        ]
        for col, val, fmt in totals:
            c = ws_main.cell(row=cur, column=col)
            c.fill = PatternFill('solid', fgColor=COL_TOTAL)
            c.border = BORDER_ALL
            if val is not None:
                c.value = val
                c.font = Font(name='Calibri', size=11, bold=True)
                c.alignment = _a_right()
                if fmt:
                    c.number_format = fmt

        ws_main.row_dimensions[cur].height = 24
        
        # Column widths for main sheet
        widths = [14, 16, 35, 14, 14, 14, 14, 14, 18, 18, 20]
        for idx, w in enumerate(widths, 1):
            ws_main.column_dimensions[get_column_letter(idx)].width = w

        # Detail Tabs (one per active WIP order)
        wip_details_by_order = {}
        for d in wip_details:
            order_number = d.get('OrderNumber')
            if order_number not in wip_details_by_order:
                wip_details_by_order[order_number] = []
            wip_details_by_order[order_number].append(d)

        for order_num, boards in wip_details_by_order.items():
            sheet_title = str(order_num)[:30]
            ws_detail = wb.create_sheet(title=sheet_title)
            self._build_wip_detail_sheet(ws_detail, order_num, boards, now)

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    def _build_wip_detail_sheet(self, ws: Worksheet, order_num: str, boards: List[dict], now: datetime) -> None:
        # 1) Title + Subtitle
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
        title = ws.cell(row=1, column=1, value=f"Dettaglio WIP Ordine {order_num}")
        title.font = Font(name='Calibri', size=13, bold=True, color='FFFFFF')
        title.fill = PatternFill('solid', fgColor=COL_TITLE)
        title.alignment = _a_center()
        ws.row_dimensions[1].height = 24

        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5)
        sub = ws.cell(row=2, column=1, value=f"Totale schede in lavorazione: {len(boards)} \u00b7 Generato: {now.strftime('%d/%m/%Y %H:%M')}")
        sub.font = Font(name='Calibri', size=10, italic=True, color='595959')
        sub.alignment = Alignment(horizontal='right', vertical='center')

        # 2) Headers
        header_row = 4
        headers = [
            'Board ID (IDBoard)', 'Data/Ora Ultima Scansione', 
            'Fase Corrente', 'Abbreviazione Fase', 'Stato'
        ]
        for idx, h in enumerate(headers, start=1):
            c = ws.cell(row=header_row, column=idx, value=h)
            c.font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
            c.fill = PatternFill('solid', fgColor=COL_HEADER)
            c.alignment = _a_center()
            c.border = BORDER_ALL
        ws.row_dimensions[header_row].height = 24
        ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate

        # 3) Data
        cur = header_row + 1
        for b in boards:
            ws.cell(row=cur, column=1, value=b.get('IDBoard')).alignment = _a_center()
            
            scan_time = b.get('ScanTimeStart')
            time_str = scan_time.strftime('%d/%m/%Y %H:%M:%S') if isinstance(scan_time, datetime) else str(scan_time)
            ws.cell(row=cur, column=2, value=time_str).alignment = _a_center()
            
            ws.cell(row=cur, column=3, value=b.get('PhaseName')).alignment = _a_left()
            ws.cell(row=cur, column=4, value=b.get('PhaseAbbreviation')).alignment = _a_center()
            
            is_pass = b.get('IsPass')
            status_text = 'OK' if is_pass == 1 else 'FAIL'
            status_cell = ws.cell(row=cur, column=5, value=status_text)
            status_cell.alignment = _a_center()
            
            if is_pass == 1:
                status_cell.fill = PatternFill('solid', fgColor='C6EFCE')
                status_cell.font = Font(name='Calibri', size=10, bold=True, color='006100')
            else:
                status_cell.fill = PatternFill('solid', fgColor='FFC7CE')
                status_cell.font = Font(name='Calibri', size=10, bold=True, color='9C0006')

            for col in range(1, 5):
                ws.cell(row=cur, column=col).font = Font(name='Calibri', size=10)
                ws.cell(row=cur, column=col).border = BORDER_ALL
            ws.cell(row=cur, column=5).border = BORDER_ALL
            cur += 1

        widths = [20, 24, 30, 16, 12]
        for idx, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(idx)].width = w

    def _write_wip_note(self, ws: Worksheet, start_row: int, text: str) -> None:
        ws.merge_cells(
            start_row=start_row, start_column=1,
            end_row=start_row + 1, end_column=11,
        )
        c = ws.cell(row=start_row, column=1, value=text)
        c.fill = PatternFill('solid', fgColor=COL_NOTE_BG)
        thick = Side(style='medium', color=COL_NOTE_BORDER)
        c.border = Border(left=thick, right=thick, top=thick, bottom=thick)
        c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=1)
        c.font = Font(name='Calibri', size=10, italic=True, color='333333')
        ws.row_dimensions[start_row].height = 20
        ws.row_dimensions[start_row + 1].height = 20

    # ------------------------------------------------------------- Main sheet
    def _build_main_sheet(
        self,
        ws: Worksheet,
        now: datetime,
        prod_day: date,
        year: int,
        month: int,
        metrics: dict,
        rows: List[Tuple[str, date, int]],
        details: Dict[str, Dict],
        price_map: Dict[str, float],
        sl_orders: set,
    ) -> None:
        has_semilavorati = bool(sl_orders)
        last_col = self.LAST_COL
        month_label = f"{_IT_MONTHS[month - 1]} {year}"

        # 1) Title + subtitle
        self._write_title(ws, row=1, text=f"Produzione Dettagliata \u2014 {month_label}")
        self._write_subtitle(
            ws,
            row=2,
            text=(
                f"Generato il {now.strftime('%d/%m/%Y %H:%M')} \u00b7 "
                f"Giorno produttivo corrente: {prod_day.strftime('%d/%m/%Y')}"
            ),
        )

        # 2) KPI block
        cur = 4
        cur = self._write_kpi_block(ws, start_row=cur, metrics=metrics)

        # 3) Note about semilavorati
        cur += 1
        if has_semilavorati:
            cur = self._write_note(
                ws, start_row=cur,
                text=(
                    "Nota: i codici prodotto che iniziano con 'SL' sono semilavorati, "
                    "evidenziati qui in grassetto e corsivo. Normalmente non vengono "
                    "valorizzati con un prezzo di vendita, perche' verranno inglobati "
                    "in un prodotto finito ('PF'); il loro valore unitario puo' quindi "
                    "risultare pari a zero. Nel tab 'Semilavorati' e' riportato il "
                    "valore di sola manodopera recuperato dal vecchio ERP."
                ),
            )
            cur += 1

        # 4) Data table
        header_row = cur
        self._write_data_headers(
            ws,
            row=header_row,
            headers=[
                'Settimana', 'Giorno', 'Order Number', 'Product Code',
                'Product Name', 'Quantita\' Prodotta', 'Prezzo Unitario', 'Valore Prodotto',
            ],
        )
        ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate

        # Build structure: week -> day -> [(order, qty), ...]
        rows_by_day: Dict[date, List[Tuple[str, int]]] = {}
        for order, pd, qty in rows:
            rows_by_day.setdefault(pd, []).append((order, qty))
        sorted_days = sorted(rows_by_day.keys())

        cur = header_row + 1
        grand_qty = 0
        grand_value = 0.0

        if not sorted_days:
            self._write_empty_row(ws, row=cur, text='Nessuna produzione registrata nel mese corrente.')
            cur += 1
        else:
            def iso_key(d: date) -> Tuple[int, int]:
                iso = d.isocalendar()
                return (iso[0], iso[1])

            for (_iso_year, iso_week), day_iter in groupby(sorted_days, key=iso_key):
                week_days = list(day_iter)
                week_qty = 0
                week_value = 0.0

                # Week header row (level 0)
                week_header_row = cur
                self._write_week_header(
                    ws, row=week_header_row, iso_week=iso_week, week_days=week_days,
                )
                cur += 1

                for d in week_days:
                    day_qty = 0
                    day_value = 0.0

                    # Day header row (level 1)
                    day_header_row = cur
                    cur += 1

                    # Detail rows (level 2)
                    detail_start = cur
                    orders_for_day = sorted(rows_by_day[d], key=lambda x: x[0])
                    for order, qty in orders_for_day:
                        det = details.get(order, {})
                        price = float(price_map.get(order, 0.0) or 0.0)
                        value = qty * price
                        is_sl = order in sl_orders

                        self._write_detail_row(
                            ws, row=cur,
                            week_label=f"W{iso_week}",
                            day_label=self._day_label(d),
                            order=order,
                            product_code=det.get('productCode', ''),
                            product_name=det.get('productName', ''),
                            qty=int(qty),
                            price=price,
                            value=value,
                            is_sl=is_sl,
                        )
                        ws.row_dimensions[cur].outline_level = 2
                        day_qty += int(qty)
                        day_value += value
                        cur += 1

                    # Write day header row retroactively (with day totals)
                    self._write_day_header(
                        ws, row=day_header_row,
                        day=d,
                        qty=day_qty,
                        value=day_value,
                    )
                    ws.row_dimensions[day_header_row].outline_level = 1

                    week_qty += day_qty
                    week_value += day_value

                # Week subtotal (level 0)
                self._write_week_subtotal(ws, row=cur, iso_week=iso_week, qty=week_qty, value=week_value)
                cur += 1

                grand_qty += week_qty
                grand_value += week_value

            # Grand total (blank separator then total)
            cur += 1
            self._write_grand_total(ws, row=cur, qty=grand_qty, value=grand_value)
            cur += 1

        # Column widths
        widths = [11, 16, 16, 16, 44, 16, 16, 18]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # Outline pref: summary above (headers visible when collapsed)
        ws.sheet_properties.outlinePr.summaryBelow = False
        ws.sheet_properties.outlinePr.summaryRight = False

        self._apply_print_setup(ws)

    # --------------------------------------------------------- Semilavorati
    def _build_semilavorati_sheet(
        self,
        ws: Worksheet,
        now: datetime,
        prod_day: date,
        year: int,
        month: int,
        rows: List[Tuple[str, date, int]],
        details: Dict[str, Dict],
        labor_price_map: Dict[str, float],
    ) -> None:
        last_col = self.LAST_COL
        month_label = f"{_IT_MONTHS[month - 1]} {year}"

        self._write_title(ws, row=1, text=f"Semilavorati \u2014 {month_label}")
        self._write_subtitle(
            ws,
            row=2,
            text=(
                f"Generato il {now.strftime('%d/%m/%Y %H:%M')} \u00b7 "
                f"Giorno produttivo corrente: {prod_day.strftime('%d/%m/%Y')}"
            ),
        )

        # Note about labor-only valorization. "sola manodopera" is underlined.
        ws.merge_cells(start_row=4, start_column=1, end_row=5, end_column=last_col)
        c = ws.cell(row=4, column=1)
        c.fill = PatternFill('solid', fgColor=COL_NOTE_BG)
        thick = Side(style='medium', color=COL_NOTE_BORDER)
        c.border = Border(left=thick, right=thick, top=thick, bottom=thick)
        c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=1)

        # openpyxl supports rich text via CellRichText (openpyxl>=3.1) to underline
        # only "SOLA MANODOPERA". We fall back to plain text with explicit markers
        # if rich text is not available.
        try:
            from openpyxl.cell.rich_text import CellRichText, TextBlock
            from openpyxl.cell.text import InlineFont
            normal = InlineFont(rFont='Calibri', sz=10, color='333333')
            emph = InlineFont(rFont='Calibri', sz=10, b=True, u='single', color='7F3F00')
            rt = CellRichText([
                TextBlock(normal,
                    "I codici prodotto che iniziano con 'SL' sono semilavorati e "
                    "normalmente non vengono valorizzati, in quanto verranno inglobati "
                    "in un prodotto finito ('PF'). "
                    "Il valore unitario qui riportato rappresenta "),
                TextBlock(emph, "SOLO LA MANODOPERA"),
                TextBlock(normal,
                    ", estratto dal vecchio ERP (resetservices). "
                    "Per gli ordini senza corrispondenza il valore e' zero."),
            ])
            c.value = rt
        except Exception:
            c.value = (
                "I codici prodotto che iniziano con 'SL' sono semilavorati e normalmente "
                "non vengono valorizzati (saranno inglobati in un 'PF'). Il valore "
                "unitario qui riportato rappresenta SOLO LA MANODOPERA, estratto "
                "dal vecchio ERP (resetservices). Per gli ordini senza corrispondenza "
                "il valore e' zero."
            )
            c.font = Font(name='Calibri', size=10, color='333333')

        ws.row_dimensions[4].height = 26
        ws.row_dimensions[5].height = 26

        # Headers
        header_row = 7
        self._write_data_headers(
            ws, row=header_row,
            headers=[
                'Settimana', 'Giorno', 'Order Number', 'Product Code',
                'Product Name', 'Quantita\' Prodotta',
                'Prezzo Unit. (manodopera)', 'Valore (manodopera)',
            ],
        )
        ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate

        # Rows grouped by week/day like main sheet but single outline level by day
        rows_by_day: Dict[date, List[Tuple[str, int]]] = {}
        for order, pd, qty in rows:
            rows_by_day.setdefault(pd, []).append((order, qty))
        sorted_days = sorted(rows_by_day.keys())

        cur = header_row + 1
        grand_qty = 0
        grand_value = 0.0

        if not sorted_days:
            self._write_empty_row(ws, row=cur, text='Nessun semilavorato prodotto nel mese corrente.')
            cur += 1
        else:
            def iso_key(d: date) -> Tuple[int, int]:
                iso = d.isocalendar()
                return (iso[0], iso[1])

            for (_iso_year, iso_week), day_iter in groupby(sorted_days, key=iso_key):
                week_days = list(day_iter)
                week_qty = 0
                week_value = 0.0
                self._write_week_header(ws, row=cur, iso_week=iso_week, week_days=week_days)
                cur += 1

                for d in week_days:
                    day_qty = 0
                    day_value = 0.0
                    day_header_row = cur
                    cur += 1

                    for order, qty in sorted(rows_by_day[d], key=lambda x: x[0]):
                        det = details.get(order, {})
                        price = float(labor_price_map.get(order, 0.0) or 0.0)
                        value = qty * price

                        self._write_detail_row(
                            ws, row=cur,
                            week_label=f"W{iso_week}",
                            day_label=self._day_label(d),
                            order=order,
                            product_code=det.get('productCode', ''),
                            product_name=det.get('productName', ''),
                            qty=int(qty),
                            price=price,
                            value=value,
                        )
                        ws.row_dimensions[cur].outline_level = 2
                        day_qty += int(qty)
                        day_value += value
                        cur += 1

                    self._write_day_header(ws, row=day_header_row, day=d, qty=day_qty, value=day_value)
                    ws.row_dimensions[day_header_row].outline_level = 1
                    week_qty += day_qty
                    week_value += day_value

                self._write_week_subtotal(ws, row=cur, iso_week=iso_week, qty=week_qty, value=week_value)
                cur += 1
                grand_qty += week_qty
                grand_value += week_value

            cur += 1
            self._write_grand_total(ws, row=cur, qty=grand_qty, value=grand_value, label='TOTALE MANODOPERA')
            cur += 1

        widths = [11, 16, 16, 16, 42, 16, 20, 20]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.sheet_properties.outlinePr.summaryBelow = False
        ws.sheet_properties.outlinePr.summaryRight = False

        self._apply_print_setup(ws)

    # --------------------------------------------------------------- Helpers
    @staticmethod
    def _day_label(d: date) -> str:
        return f"{_IT_WEEKDAYS[d.weekday()]} {d.strftime('%d/%m')}"

    def _write_title(self, ws: Worksheet, row: int, text: str) -> None:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=self.LAST_COL)
        c = ws.cell(row=row, column=1, value=text)
        c.font = Font(name='Calibri', size=15, bold=True, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor=COL_TITLE)
        c.alignment = _a_center()
        ws.row_dimensions[row].height = 30

    def _write_subtitle(self, ws: Worksheet, row: int, text: str) -> None:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=self.LAST_COL)
        c = ws.cell(row=row, column=1, value=text)
        c.font = Font(name='Calibri', size=10, italic=True, color='595959')
        c.alignment = Alignment(horizontal='right', vertical='center')

    def _write_kpi_block(self, ws: Worksheet, start_row: int, metrics: dict) -> int:
        """Scrive il blocco KPI a 4 colonne (label-valore ripetuto due volte)."""
        # Band header
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=self.LAST_COL)
        c = ws.cell(row=start_row, column=1, value='Situazione KPI e Previsioni')
        c.font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor=COL_KPI_BAND)
        c.alignment = _a_center()
        ws.row_dimensions[start_row].height = 22

        # 2 colonne: A-B label, C-D valore | E-F label, G-H valore
        kpi_rows = [
            ('Target Giornaliero', metrics.get('dailyTarget'),
             'Target Mensile', metrics.get('monthlyTarget')),
            ('Valore Prodotto Oggi', metrics.get('todayValue'),
             'Valore Prodotto Mese', metrics.get('monthValue')),
            ('Gap Giornaliero', metrics.get('dailyGap'),
             'Gap Mensile', metrics.get('monthlyGap')),
            ('Previsione Fine Giorno', metrics.get('forecastDay'),
             'Previsione Fine Mese', metrics.get('forecastMonth')),
            ('Aggiustamento Richiesto', metrics.get('requiredDailyAdjustment'),
             'Giorni Lavorativi Mese',
             f"{metrics.get('workingDaysInMonth', 0)}  "
             f"(residui: {metrics.get('remainingWorkingDays', 0)})"),
        ]

        r = start_row + 1
        for (la, va, lb, vb) in kpi_rows:
            self._write_kpi_cell(ws, row=r, col_label=1, col_value=3, label=la, value=va)
            self._write_kpi_cell(ws, row=r, col_label=5, col_value=7, label=lb, value=vb)
            ws.row_dimensions[r].height = 20
            r += 1

        return r - 1

    def _write_kpi_cell(
        self, ws: Worksheet, row: int, col_label: int, col_value: int, label: str, value,
    ) -> None:
        # Label: merge col_label .. col_label+1
        ws.merge_cells(start_row=row, start_column=col_label, end_row=row, end_column=col_label + 1)
        lc = ws.cell(row=row, column=col_label, value=label)
        lc.font = Font(name='Calibri', size=10, bold=True, color='1F4E78')
        lc.fill = PatternFill('solid', fgColor=COL_KPI_LABEL)
        lc.alignment = _a_left(wrap=False, indent=1)
        lc.border = BORDER_ALL

        # Value: merge col_value .. col_value+1
        ws.merge_cells(start_row=row, start_column=col_value, end_row=row, end_column=col_value + 1)
        vc = ws.cell(row=row, column=col_value)
        vc.alignment = _a_right()
        vc.border = BORDER_ALL
        vc.font = Font(name='Calibri', size=11, bold=True, color='1F4E78')
        vc.fill = PatternFill('solid', fgColor=COL_KPI_VALUE)

        if isinstance(value, (int, float)):
            vc.value = float(value)
            vc.number_format = CURRENCY_FMT
        else:
            vc.value = value

        # propagate border to merged-in cells
        for col in (col_label + 1, col_value + 1):
            ws.cell(row=row, column=col).border = BORDER_ALL

    def _write_note(self, ws: Worksheet, start_row: int, text: str) -> int:
        ws.merge_cells(
            start_row=start_row, start_column=1,
            end_row=start_row + 1, end_column=self.LAST_COL,
        )
        c = ws.cell(row=start_row, column=1, value=text)
        c.fill = PatternFill('solid', fgColor=COL_NOTE_BG)
        thick = Side(style='medium', color=COL_NOTE_BORDER)
        c.border = Border(left=thick, right=thick, top=thick, bottom=thick)
        c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=1)
        c.font = Font(name='Calibri', size=10, italic=True, color='333333')
        ws.row_dimensions[start_row].height = 22
        ws.row_dimensions[start_row + 1].height = 22
        return start_row + 1

    def _write_data_headers(self, ws: Worksheet, row: int, headers: List[str]) -> None:
        for idx, h in enumerate(headers, start=1):
            c = ws.cell(row=row, column=idx, value=h)
            c.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
            c.fill = PatternFill('solid', fgColor=COL_HEADER)
            c.alignment = _a_center(wrap=True)
            c.border = BORDER_ALL
        ws.row_dimensions[row].height = 30

    def _write_empty_row(self, ws: Worksheet, row: int, text: str) -> None:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=self.LAST_COL)
        c = ws.cell(row=row, column=1, value=text)
        c.font = Font(name='Calibri', size=10, italic=True, color='595959')
        c.alignment = _a_center()

    def _write_week_header(self, ws: Worksheet, row: int, iso_week: int, week_days: List[date]) -> None:
        week_range = f"{week_days[0].strftime('%d/%m')} \u2013 {week_days[-1].strftime('%d/%m')}"
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=self.LAST_COL)
        c = ws.cell(row=row, column=1, value=f"Settimana {iso_week}  ({week_range})")
        c.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor=COL_WEEK)
        c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        ws.row_dimensions[row].height = 22

    def _write_day_header(self, ws: Worksheet, row: int, day: date, qty: int, value: float) -> None:
        # Columns: A-E = day label + descrizione, F=qty, G=(blank), H=value
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        c = ws.cell(row=row, column=1, value=f"  {self._day_label(day)}")
        c.font = Font(name='Calibri', size=10, bold=True, color='1F4E78')
        c.fill = PatternFill('solid', fgColor=COL_DAY)
        c.alignment = Alignment(horizontal='left', vertical='center', indent=2)

        for col in range(1, 6):
            ws.cell(row=row, column=col).fill = PatternFill('solid', fgColor=COL_DAY)
            ws.cell(row=row, column=col).border = BORDER_ALL

        qc = ws.cell(row=row, column=6, value=int(qty))
        qc.font = Font(name='Calibri', size=10, bold=True, color='1F4E78')
        qc.fill = PatternFill('solid', fgColor=COL_DAY)
        qc.alignment = _a_right()
        qc.number_format = INT_FMT
        qc.border = BORDER_ALL

        ws.cell(row=row, column=7).fill = PatternFill('solid', fgColor=COL_DAY)
        ws.cell(row=row, column=7).border = BORDER_ALL

        vc = ws.cell(row=row, column=8, value=float(value))
        vc.font = Font(name='Calibri', size=10, bold=True, color='1F4E78')
        vc.fill = PatternFill('solid', fgColor=COL_DAY)
        vc.alignment = _a_right()
        vc.number_format = CURRENCY_FMT
        vc.border = BORDER_ALL

    def _write_detail_row(
        self,
        ws: Worksheet, row: int,
        week_label: str, day_label: str, order: str,
        product_code: str, product_name: str,
        qty: int, price: float, value: float,
        is_sl: bool = False,
    ) -> None:
        # Semilavorati evidenziati in grassetto + corsivo (colore bruno-rossiccio per distinguerli)
        if is_sl:
            base_font = Font(name='Calibri', size=10, bold=True, italic=True, color='7F3F00')
            num_font = Font(name='Calibri', size=10, bold=True, italic=True, color='7F3F00')
        else:
            base_font = Font(name='Calibri', size=10)
            num_font = Font(name='Calibri', size=10)

        cells = [
            (1, week_label, _a_center()),
            (2, day_label, _a_center()),
            (3, order, _a_center()),
            (4, product_code, _a_center()),
            (5, product_name, _a_left()),
        ]
        for col, val, align in cells:
            c = ws.cell(row=row, column=col, value=val)
            c.alignment = align
            c.font = base_font
            c.border = BORDER_ALL

        qc = ws.cell(row=row, column=6, value=int(qty))
        qc.alignment = _a_right()
        qc.number_format = INT_FMT
        qc.font = num_font
        qc.border = BORDER_ALL

        pc = ws.cell(row=row, column=7, value=float(price))
        pc.alignment = _a_right()
        pc.number_format = CURRENCY_FMT
        pc.font = num_font
        pc.border = BORDER_ALL

        vc = ws.cell(row=row, column=8, value=float(value))
        vc.alignment = _a_right()
        vc.number_format = CURRENCY_FMT
        vc.font = num_font
        vc.border = BORDER_ALL

    def _write_week_subtotal(self, ws: Worksheet, row: int, iso_week: int, qty: int, value: float) -> None:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        c = ws.cell(row=row, column=1, value=f"Totale settimana {iso_week}")
        c.font = Font(name='Calibri', size=10, bold=True, color='1F4E78')
        c.alignment = Alignment(horizontal='right', vertical='center', indent=1)
        for col in range(1, 6):
            ws.cell(row=row, column=col).fill = PatternFill('solid', fgColor=COL_SUBTOTAL)
            ws.cell(row=row, column=col).border = BORDER_ALL

        qc = ws.cell(row=row, column=6, value=int(qty))
        qc.font = Font(name='Calibri', size=10, bold=True)
        qc.fill = PatternFill('solid', fgColor=COL_SUBTOTAL)
        qc.alignment = _a_right()
        qc.number_format = INT_FMT
        qc.border = BORDER_ALL

        ws.cell(row=row, column=7).fill = PatternFill('solid', fgColor=COL_SUBTOTAL)
        ws.cell(row=row, column=7).border = BORDER_ALL

        vc = ws.cell(row=row, column=8, value=float(value))
        vc.font = Font(name='Calibri', size=10, bold=True)
        vc.fill = PatternFill('solid', fgColor=COL_SUBTOTAL)
        vc.alignment = _a_right()
        vc.number_format = CURRENCY_FMT
        vc.border = BORDER_ALL

    def _write_grand_total(
        self, ws: Worksheet, row: int, qty: int, value: float, label: str = 'TOTALE MESE',
    ) -> None:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        c = ws.cell(row=row, column=1, value=label)
        c.font = Font(name='Calibri', size=13, bold=True)
        c.alignment = Alignment(horizontal='right', vertical='center', indent=1)
        for col in range(1, 6):
            ws.cell(row=row, column=col).fill = PatternFill('solid', fgColor=COL_TOTAL)
            ws.cell(row=row, column=col).border = BORDER_ALL

        qt = ws.cell(row=row, column=6, value=int(qty))
        qt.font = Font(name='Calibri', size=12, bold=True)
        qt.fill = PatternFill('solid', fgColor=COL_TOTAL)
        qt.alignment = _a_right()
        qt.number_format = INT_FMT
        qt.border = BORDER_ALL

        ws.cell(row=row, column=7).fill = PatternFill('solid', fgColor=COL_TOTAL)
        ws.cell(row=row, column=7).border = BORDER_ALL

        vt = ws.cell(row=row, column=8, value=float(value))
        vt.font = Font(name='Calibri', size=12, bold=True)
        vt.fill = PatternFill('solid', fgColor=COL_TOTAL)
        vt.alignment = _a_right()
        vt.number_format = CURRENCY_FMT
        vt.border = BORDER_ALL

        ws.row_dimensions[row].height = 26

    def _apply_print_setup(self, ws: Worksheet) -> None:
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.print_options.horizontalCentered = True
        ws.page_margins.left = 0.3
        ws.page_margins.right = 0.3
        ws.page_margins.top = 0.4
        ws.page_margins.bottom = 0.4
